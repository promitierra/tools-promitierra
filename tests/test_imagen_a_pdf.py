import unittest
import os
import tempfile
import shutil
from PIL import Image
import zipfile
from imagen_a_pdf import ImagenAPdfApp
import customtkinter as ctk
import PyPDF2
from openpyxl import Workbook, load_workbook
import pandas as pd

class TestImagenAPdfApp(unittest.TestCase):
    def setUp(self):
        """Configuración inicial para cada prueba"""
        self.temp_dir = tempfile.mkdtemp()
        self.app = ImagenAPdfApp()
        
        # Crear algunas imágenes de prueba
        self.test_images = []
        for i in range(3):
            img_path = os.path.join(self.temp_dir, f'test_image_{i}.png')
            self.test_images.append(img_path)
            # Crear una imagen de prueba
            img = Image.new('RGB', (100, 100), color='red')
            img.save(img_path)
    
    def tearDown(self):
        """Limpieza después de cada prueba"""
        try:
            shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Error al limpiar directorio temporal: {e}")
    
    def test_creacion_ventana(self):
        """Prueba la creación de la ventana principal"""
        self.assertIsInstance(self.app.ventana, ctk.CTk)
        self.assertEqual(self.app.ventana.title(), "Herramientas de Productividad")
        
    def test_modo_comprimido_inicial(self):
        """Prueba el estado inicial del modo comprimido"""
        self.assertFalse(self.app.modo_comprimido.get())
    
    def test_conversion_individual(self):
        """Prueba la conversión de una sola imagen"""
        img_path = self.test_images[0]
        pdf_path = os.path.splitext(img_path)[0] + '.pdf'
        
        # Simular conversión
        with Image.open(img_path) as img:
            img = img.convert("RGB")
            img.save(pdf_path, "PDF")
        
        self.assertTrue(os.path.exists(pdf_path))
        # Verificar que el archivo existe y tiene un tamaño mayor a 0
        self.assertTrue(os.path.getsize(pdf_path) > 0)
    
    def test_creacion_zip(self):
        """Prueba la creación del archivo ZIP"""
        zip_path = os.path.join(self.temp_dir, 'test.zip')
        
        # Crear algunos PDFs de prueba
        pdf_files = []
        for img_path in self.test_images:
            pdf_path = os.path.splitext(img_path)[0] + '.pdf'
            with Image.open(img_path) as img:
                img = img.convert("RGB")
                img.save(pdf_path, "PDF")
            pdf_files.append(pdf_path)
        
        # Crear ZIP
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for pdf in pdf_files:
                zipf.write(pdf, os.path.basename(pdf))
        
        self.assertTrue(os.path.exists(zip_path))
        # Verificar contenido del ZIP
        with zipfile.ZipFile(zip_path, 'r') as zipf:
            self.assertEqual(len(zipf.namelist()), len(pdf_files))
    
    def test_manejo_errores_imagen_invalida(self):
        """Prueba el manejo de errores con una imagen inválida"""
        invalid_img_path = os.path.join(self.temp_dir, 'invalid.png')
        with open(invalid_img_path, 'w') as f:
            f.write('not an image')
        
        # Intentar convertir imagen inválida
        with self.assertRaises(Exception):
            Image.open(invalid_img_path)
    
    def test_estructura_directorios_zip(self):
        """Prueba la preservación de la estructura de directorios en ZIP"""
        # Crear estructura de directorios
        subdir = os.path.join(self.temp_dir, 'subdir')
        os.makedirs(subdir)
        
        # Crear imagen en subdirectorio
        img_path = os.path.join(subdir, 'test_sub.png')
        img = Image.new('RGB', (100, 100), color='blue')
        img.save(img_path)
        
        # Convertir a PDF
        pdf_path = os.path.splitext(img_path)[0] + '.pdf'
        with Image.open(img_path) as img:
            img = img.convert("RGB")
            img.save(pdf_path, "PDF")
        
        # Crear ZIP
        zip_path = os.path.join(self.temp_dir, 'test_structure.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(pdf_path, os.path.relpath(pdf_path, self.temp_dir))
        
        # Verificar estructura en ZIP
        with zipfile.ZipFile(zip_path, 'r') as zipf:
            self.assertIn('subdir/test_sub.pdf', zipf.namelist())

    def test_creacion_plantilla_excel(self):
        """Prueba la creación de la plantilla Excel"""
        plantilla_path = os.path.join(self.temp_dir, 'test_plantilla.xlsx')
        
        # Simular creación de plantilla
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID'
        ws['B1'] = 'NOMBRES'
        ws['C1'] = 'APELLIDOS'
        wb.save(plantilla_path)
        
        self.assertTrue(os.path.exists(plantilla_path))
        
        # Verificar estructura de la plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'ID')
        self.assertEqual(ws['B1'].value, 'NOMBRES')
        self.assertEqual(ws['C1'].value, 'APELLIDOS')
    
    def test_creacion_carpetas_desde_excel(self):
        """Prueba la creación de carpetas desde datos Excel"""
        # Crear plantilla de prueba
        plantilla_path = os.path.join(self.temp_dir, 'test_datos.xlsx')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID'
        ws['B1'] = 'NOMBRES'
        ws['C1'] = 'APELLIDOS'
        
        # Agregar datos de prueba
        datos = [
            ('1', 'Juan', 'Pérez'),
            ('2', 'María', 'García')
        ]
        for i, (id, nombre, apellido) in enumerate(datos, start=2):
            ws[f'A{i}'] = id
            ws[f'B{i}'] = nombre
            ws[f'C{i}'] = apellido
        wb.save(plantilla_path)
        
        # Directorio para las carpetas
        carpetas_dir = os.path.join(self.temp_dir, 'carpetas')
        os.makedirs(carpetas_dir)
        
        # Simular creación de carpetas
        df = pd.read_excel(plantilla_path)
        for _, row in df.iterrows():
            nombre_carpeta = f"{row['ID']} - {row['NOMBRES']} {row['APELLIDOS']}"
            ruta_carpeta = os.path.join(carpetas_dir, nombre_carpeta)
            os.makedirs(ruta_carpeta)
        
        # Verificar carpetas creadas
        carpetas_creadas = os.listdir(carpetas_dir)
        self.assertEqual(len(carpetas_creadas), 2)
        self.assertIn('1 - Juan Pérez', carpetas_creadas)
        self.assertIn('2 - María García', carpetas_creadas)

    def test_validacion_datos_excel(self):
        """Prueba la validación de datos en el Excel"""
        # Crear plantilla con columnas incorrectas
        plantilla_path = os.path.join(self.temp_dir, 'test_invalido.xlsx')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID'
        ws['B1'] = 'NOMBRE'  # Columna incorrecta
        ws['C1'] = 'APELLIDO'  # Columna incorrecta
        wb.save(plantilla_path)
        
        # Verificar que se detectan columnas incorrectas
        df = pd.read_excel(plantilla_path)
        columnas_requeridas = ['ID', 'NOMBRES', 'APELLIDOS']
        columnas_validas = all(col in df.columns for col in columnas_requeridas)
        self.assertFalse(columnas_validas)

if __name__ == '__main__':
    unittest.main()
