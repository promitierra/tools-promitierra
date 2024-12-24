print("Iniciando aplicación...")

import os
import customtkinter as ctk
from PIL import Image
from tkinter import filedialog, messagebox
import threading
import zipfile
from datetime import datetime
import shutil
import tempfile
import pandas as pd
import openpyxl
from openpyxl import Workbook
import re

class ImagenAPdfApp:
    def __init__(self):
        # Configuración de la ventana principal
        self.ventana = ctk.CTk()
        self.ventana.title("Herramientas de Productividad")
        self.ventana.geometry("500x500")  # Reducimos significativamente la altura
        self.ventana.resizable(False, False)  # Bloqueamos el redimensionamiento
        
        # Configurar el tema
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")
        
        # Variables de control
        self.procesando = False
        self.modo_comprimido = ctk.BooleanVar(value=False)
        self.directorio_salida = None
        
        # Crear el contenido de la ventana
        self.crear_widgets()

    def iniciar(self):
        # Iniciar el mainloop
        self.ventana.mainloop()

    def crear_widgets(self):
        # Preparar la estructura para agregar pestañas
        self.notebook = ctk.CTkTabview(self.ventana)
        self.notebook.pack(pady=20, padx=20, fill="both", expand=True)
        
        # Agregar pestañas en el orden deseado
        self.pestaña_carpetas = self.notebook.add("Crear Carpetas")  # Primera pestaña
        self.pestaña_principal = self.notebook.add("imagenes a PDFs")     # Segunda pestaña
        
        # Crear contenido de las pestañas
        self.crear_contenido_pestaña_carpetas()
        self.crear_contenido_pestaña_principal()
        

    def crear_contenido_pestaña_carpetas(self):
        # Frame principal con padding mínimo
        frame = ctk.CTkFrame(self.pestaña_carpetas)
        frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        # Título
        titulo = ctk.CTkLabel(frame, text="Creación Masiva de Carpetas", 
                            font=ctk.CTkFont(size=20, weight="bold"))
        titulo.pack(pady=5)

        # Descripción
        descripcion1 = ctk.CTkLabel(frame, 
            text="Descarga la plantilla Excel, completa los datos con ID, NOMBRES y APELLIDOS.",
            font=ctk.CTkFont(size=12),
            text_color="#CCCCCC"
        )
        descripcion1.pack(pady=2)

        descripcion2 = ctk.CTkLabel(frame, 
            text="Luego, carga el archivo para crear automáticamente las carpetas.",
            font=ctk.CTkFont(size=12),
            text_color="#CCCCCC"
        )
        descripcion2.pack(pady=(0,5))
        
        # Frame para la primera fila (plantilla y carga)
        fila1_frame = ctk.CTkFrame(frame)
        fila1_frame.pack(fill="x", padx=10, pady=10)
        
        # Columna izquierda: Descargar plantilla
        btn_descargar = ctk.CTkButton(
            fila1_frame,
            text="📥 Descargar Plantilla",
            command=self.descargar_plantilla
        )
        btn_descargar.pack(side="left", padx=5)
        
        # Columna derecha: Cargar plantilla
        btn_cargar = ctk.CTkButton(
            fila1_frame,
            text="📤 Cargar Plantilla",
            command=self.cargar_plantilla
        )
        btn_cargar.pack(side="right", padx=5)
        
        # Etiqueta de estado
        self.estado_carpetas_label = ctk.CTkLabel(frame, 
            text="Estado: Esperando plantilla...",
            font=ctk.CTkFont(size=12)
        )
        self.estado_carpetas_label.pack(pady=10)
        
        # Área de detalles para la creación de carpetas
        self.detalles_carpetas_text = ctk.CTkTextbox(frame, height=100)
        self.detalles_carpetas_text.pack(fill="x", padx=10, pady=(5,10))
        self.detalles_carpetas_text.configure(state="disabled")

        # Agregar footer
        self.crear_footer(frame)
        
    def crear_contenido_pestaña_principal(self):
        # Frame principal con padding mínimo
        frame = ctk.CTkFrame(self.pestaña_principal)
        frame.pack(pady=10, padx=10, fill="both", expand=True)

        # Título
        titulo = ctk.CTkLabel(frame, text="Conversor de Imágenes a PDF", 
                            font=ctk.CTkFont(size=20, weight="bold"))
        titulo.pack(pady=5)

        # Variable para el filtro (oculta para el usuario)
        self.filtro = "*"

        # Checkbox para comprimir
        self.cb_comprimir = ctk.CTkCheckBox(
            frame,
            text="Generar archivos PDFs en un nuevo archivo ZIP",
            variable=self.modo_comprimido,
            onvalue=True,
            offvalue=False
        )
        self.cb_comprimir.pack(pady=10)

        # Botón para seleccionar carpeta
        self.btn_seleccionar = ctk.CTkButton(
            frame,
            text="Seleccionar Carpeta",
            command=self.seleccionar_carpeta
        )
        self.btn_seleccionar.pack(pady=10)

        # Barra de progreso
        self.barra_progreso = ctk.CTkProgressBar(frame)
        self.barra_progreso.pack(fill="x", padx=10, pady=5)
        self.barra_progreso.set(0)

        # Etiqueta de progreso
        self.progreso_label = ctk.CTkLabel(frame, text="0%")
        self.progreso_label.pack(pady=5)

        # Etiqueta de estado
        self.estado_label = ctk.CTkLabel(
            frame,
            text="Estado: Esperando selección de carpeta...",
            font=ctk.CTkFont(size=12)
        )
        self.estado_label.pack(pady=5)

        # Área de detalles
        self.detalles_text = ctk.CTkTextbox(frame, height=75)
        self.detalles_text.pack(fill="x", padx=10, pady=(5,10))
        self.detalles_text.configure(state="disabled")

        # Agregar footer
        self.crear_footer(frame)

    def crear_footer(self, frame_padre):
        # Frame interno para los créditos
        creditos_interno = ctk.CTkFrame(frame_padre, fg_color="transparent")
        creditos_interno.pack(side="bottom", fill="x", pady=(10, 3))

        # Primera línea: Desarrollador y año
        creditos_linea1 = ctk.CTkLabel(
            creditos_interno,
            text="Desarrollado por: Luis Fernando Moreno Montoya | 2024",
            font=ctk.CTkFont(size=13),
            text_color="#CCCCCC"
        )
        creditos_linea1.pack(pady=(10, 3))
        
        # Segunda línea: Mensaje especial (dividido en partes para colorear el corazón)
        mensaje_frame = ctk.CTkFrame(creditos_interno, fg_color="transparent")
        mensaje_frame.pack(pady=(3, 10))

        # Primera parte del mensaje
        parte1 = ctk.CTkLabel(
            mensaje_frame,
            text="Hecho con ",
            font=ctk.CTkFont(size=13),
            text_color="#CCCCCC"
        )
        parte1.pack(side="left")

        # Corazón en rojo
        corazon = ctk.CTkLabel(
            mensaje_frame,
            text="♥",
            font=ctk.CTkFont(size=13),
            text_color="#FF0000"
        )
        corazon.pack(side="left")

        # Segunda parte del mensaje
        parte2 = ctk.CTkLabel(
            mensaje_frame,
            text=" por la productividad laboral y el cuidado del tiempo",
            font=ctk.CTkFont(size=13),
            text_color="#CCCCCC"
        )
        parte2.pack(side="left")
        
        return creditos_interno

    def descargar_plantilla(self):
        try:
            # Crear un nuevo libro de Excel
            wb = Workbook()
            ws = wb.active
            
            # Configurar encabezados
            ws['A1'] = 'ID'
            ws['B1'] = 'NOMBRES'
            ws['C1'] = 'APELLIDOS'
            
            # Pedir al usuario donde guardar la plantilla
            ruta_guardado = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar plantilla como",
                initialfile="Plantilla Nombres Carpetas"
            )
            
            if ruta_guardado:
                wb.save(ruta_guardado)
                messagebox.showinfo("Éxito", "Plantilla descargada correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear la plantilla: {str(e)}")

    def cargar_plantilla(self):
        try:
            # Pedir al usuario que seleccione el archivo Excel
            ruta_excel = filedialog.askopenfilename(
                title="Seleccionar plantilla Excel",
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            
            if not ruta_excel:
                return
                
            # Pedir al usuario el directorio donde crear las carpetas
            directorio_destino = filedialog.askdirectory(
                title="Seleccionar directorio para crear carpetas"
            )
            
            if not directorio_destino:
                return
            
            # Leer el Excel
            df = pd.read_excel(ruta_excel)
            
            # Verificar columnas requeridas
            columnas_requeridas = ['ID', 'NOMBRES', 'APELLIDOS']
            if not all(col in df.columns for col in columnas_requeridas):
                messagebox.showerror("Error", "La plantilla no tiene el formato correcto")
                return
            
            # Crear carpetas
            carpetas_creadas = 0
            self.detalles_carpetas_text.configure(state="normal")
            self.detalles_carpetas_text.delete("1.0", "end")
            
            for _, row in df.iterrows():
                try:
                    # Normalizar ID: eliminar todos los espacios
                    id_value = ''.join(str(row['ID']).strip().split())
                    
                    # Normalizar nombres y apellidos usando la función normalizar_texto
                    nombres = self.normalizar_texto(row['NOMBRES'])
                    apellidos = self.normalizar_texto(row['APELLIDOS'])
                    
                    # Crear el nombre de la carpeta con el formato correcto
                    nombre_carpeta = f"{id_value} - {nombres} {apellidos}"
                    ruta_carpeta = os.path.join(directorio_destino, nombre_carpeta)
                    
                    if not os.path.exists(ruta_carpeta):
                        os.makedirs(ruta_carpeta)
                        carpetas_creadas += 1
                        self.detalles_carpetas_text.insert("end", f"✓ Creada: {nombre_carpeta}\n")
                    else:
                        self.detalles_carpetas_text.insert("end", f"⚠ Ya existe: {nombre_carpeta}\n")
                except Exception as e:
                    self.detalles_carpetas_text.insert("end", f"✗ Error al crear carpeta: {str(e)}\n")
            
            self.detalles_carpetas_text.configure(state="disabled")
            self.estado_carpetas_label.configure(text=f"Estado: {carpetas_creadas} carpetas creadas")
            
            messagebox.showinfo("Completado", f"Se han creado {carpetas_creadas} carpetas")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar la plantilla: {str(e)}")

    def seleccionar_carpeta(self):
        if self.procesando:
            messagebox.showwarning("En proceso", "Por favor espera a que termine el proceso actual")
            return
            
        carpeta = filedialog.askdirectory(title="Selecciona la carpeta con imágenes")
        if carpeta:
            # Si es modo comprimido, pedir ubicación del ZIP
            if self.modo_comprimido.get():
                fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_zip = f"PDFs_Convertidos_{fecha_actual}.zip"
                self.directorio_salida = filedialog.asksaveasfilename(
                    defaultextension=".zip",
                    initialfile=nombre_zip,
                    title="Guardar archivo ZIP como",
                    filetypes=[("Archivo ZIP", "*.zip")]
                )
                if not self.directorio_salida:
                    return
            
            # Iniciar proceso en un hilo separado
            thread = threading.Thread(target=self.procesar_carpeta, args=(carpeta,))
            thread.daemon = True
            thread.start()
    
    def procesar_carpeta(self, directorio):
        """Procesa todas las imágenes en la carpeta"""
        temp_dir = None
        try:
            self.procesando = True
            self.btn_seleccionar.configure(state="disabled")
            self.estado_label.configure(text="Estado: Buscando imágenes...")
            
            # Crear directorio temporal si es modo comprimido
            if self.modo_comprimido.get():
                temp_dir = tempfile.mkdtemp()
            
            # Encontrar todas las imágenes
            imagenes = []
            for ruta_actual, _, archivos in os.walk(directorio):
                for archivo in archivos:
                    if archivo.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                        imagenes.append(os.path.join(ruta_actual, archivo))
            
            if not imagenes:
                messagebox.showinfo("Información", "No se encontraron imágenes en la carpeta seleccionada")
                self.reset_interfaz()
                return
            
            total_imagenes = len(imagenes)
            convertidas = 0
            errores = 0
            
            self.barra_progreso.set(0)
            self.agregar_detalle(f"Encontradas {total_imagenes} imágenes para convertir")
            
            for ruta_imagen in imagenes:
                try:
                    nombre_archivo = os.path.basename(ruta_imagen)
                    self.estado_label.configure(text=f"Estado: Procesando {nombre_archivo}")
                    
                    # Determinar ruta de salida
                    if self.modo_comprimido.get():
                        # Mantener la estructura de directorios dentro del ZIP
                        ruta_relativa = os.path.relpath(ruta_imagen, directorio)
                        ruta_pdf = os.path.join(temp_dir, os.path.splitext(ruta_relativa)[0] + ".pdf")
                        # Crear directorios necesarios
                        os.makedirs(os.path.dirname(ruta_pdf), exist_ok=True)
                    else:
                        ruta_pdf = os.path.splitext(ruta_imagen)[0] + ".pdf"
                    
                    with Image.open(ruta_imagen) as img:
                        if img.mode != 'RGB':
                            img = img.convert('RGB')
                        img.save(ruta_pdf, "PDF")
                    
                    convertidas += 1
                    progreso = (convertidas / total_imagenes)
                    self.barra_progreso.set(progreso)
                    self.progreso_label.configure(text=f"{int(progreso * 100)}%")
                    self.agregar_detalle(f"✓ Convertido: {nombre_archivo}")
                    
                except Exception as e:
                    errores += 1
                    self.agregar_detalle(f"✗ Error al convertir {nombre_archivo}: {str(e)}")
                
                self.ventana.update()
            
            # Si es modo comprimido, crear el ZIP
            if self.modo_comprimido.get() and temp_dir:
                self.estado_label.configure(text="Estado: Creando archivo ZIP...")
                self.agregar_detalle("Creando archivo ZIP...")
                
                with zipfile.ZipFile(self.directorio_salida, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root, _, files in os.walk(temp_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, temp_dir)
                            zipf.write(file_path, arcname)
                
                self.agregar_detalle(f"✓ Archivo ZIP creado en: {self.directorio_salida}")
            
            # Mostrar resumen final
            mensaje = f"Proceso completado\nImágenes convertidas: {convertidas}/{total_imagenes}"
            if errores > 0:
                mensaje += f"\nErrores encontrados: {errores}"
            if self.modo_comprimido.get():
                mensaje += f"\nArchivo ZIP creado en:\n{self.directorio_salida}"
            
            messagebox.showinfo("Completado", mensaje)
            self.estado_label.configure(text="Estado: Proceso completado")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error inesperado: {str(e)}")
        finally:
            # Limpiar directorio temporal si existe
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            self.reset_interfaz()
    
    def reset_interfaz(self):
        """Resetea la interfaz a su estado inicial"""
        self.procesando = False
        self.btn_seleccionar.configure(state="normal")
        self.ventana.update()
    
    def agregar_detalle(self, texto):
        """Agrega texto al área de detalles"""
        self.detalles_text.configure(state="normal")
        self.detalles_text.insert("end", texto + "\n")
        self.detalles_text.see("end")
        self.detalles_text.configure(state="disabled")
        self.ventana.update()

    def normalizar_texto(self, texto):
        """
        Normaliza el texto eliminando espacios extra y convirtiendo a mayúsculas.
        Formato: ID - NOMBRES APELLIDOS
        Si no hay ID, se agrega "- " al inicio
        Los IDs se limpian de caracteres especiales
        """
        if texto is None:
            return " -"
            
        texto_str = str(texto).strip()
        if texto_str == "":
            return " - "
            
        # 1. Convertir a string y dividir por el guion (-)
        partes = texto_str.split('-')
        
        # 2. Procesar el ID (primera parte)
        if len(partes) > 1:
            # Unir todas las partes numéricas del ID
            id_texto = partes[0]
            for i in range(1, len(partes)-1):
                id_texto += partes[i]
            
            # Limpiar el ID de caracteres especiales y espacios
            id_parte = ''.join(c for c in id_texto if c.isalnum()).strip()
            # Procesar la parte de nombres
            nombres_parte = ' '.join(partes[-1].split())
            # Si hay nombres, retornar el formato completo
            if nombres_parte:
                return f"{id_parte} - {nombres_parte.upper()}"
            # Si no hay nombres, solo retornar el ID con el guion
            return f"{id_parte} - "
        else:
            # Si no hay guion, tratar todo como nombres
            nombres_parte = ' '.join(partes[0].split())
            if nombres_parte:
                return f"- {nombres_parte.upper()}"
            return " - "

    def test_normalizar_texto(self):
        """Pruebas unitarias para la función normalizar_texto"""
        casos_prueba = [
            ("l  uis   FERNando", "LUIS FERNANDO"),
            ("   maria    clara   ", "MARIA CLARA"),
            ("JuAn   pAbLo", "JUAN PABLO"),
            ("ana  MARIA    PEREZ  LOPEZ", "ANA MARIA PEREZ LOPEZ"),
            ("", ""),
            ("  ", ""),
            (None, ""),
            (123, "123"),
        ]
        
        for entrada, esperado in casos_prueba:
            resultado = self.normalizar_texto(entrada)
            if resultado != esperado:
                print(f"Error en caso de prueba:")
                print(f"Entrada: '{entrada}'")
                print(f"Esperado: '{esperado}'")
                print(f"Obtenido: '{resultado}'")
                return False
        return True

if __name__ == "__main__":
    app = ImagenAPdfApp()
    app.iniciar()
